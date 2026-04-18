#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations
import json
from pathlib import Path
from collections import defaultdict
import openpyxl

REPO_ROOT = Path("/Users/fengbo/GitHub/soseki-manuscripts")
MASTER_XLSX = REPO_ROOT / "master.xlsx"
IMAGES_MASTER_XLSX = REPO_ROOT / "images_master.xlsx"
MASTER_JSON = REPO_ROOT / "master.json"
BASE_URL = "https://scholarbo.github.io/soseki-manuscripts/manifests"

MASTER_REQUIRED = ["seq", "id", "title", "source", "date_year", "date_month", "date_day"]
IMAGES_REQUIRED = ["work_id", "witness_id", "manifest_id"]

def clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def load_sheet_rows(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return headers, rows

def ensure_headers(headers, required, label):
    missing = [h for h in required if h not in headers]
    if missing:
        raise SystemExit(f"{label} 缺少必要表头: {missing}")

def manifest_url(manifest_id: str) -> str:
    manifest_id = clean(manifest_id)
    if not manifest_id:
        return ""
    if manifest_id.endswith(".json"):
        return f"{BASE_URL}/{manifest_id}"
    return f"{BASE_URL}/{manifest_id}.json"

def sort_witness_key(w):
    w = clean(w)
    if w.startswith("w") and w[1:].isdigit():
        return (0, int(w[1:]))
    return (1, w)

def sort_manifest_key(url: str):
    name = url.rsplit("/", 1)[-1].replace(".json", "")
    return sort_witness_key(name.rsplit("-", 1)[-1] if "-w" in name else name)

def main():
    if not MASTER_XLSX.exists():
        raise SystemExit(f"未找到: {MASTER_XLSX}")
    if not IMAGES_MASTER_XLSX.exists():
        raise SystemExit(f"未找到: {IMAGES_MASTER_XLSX}")

    master_headers, master_rows = load_sheet_rows(MASTER_XLSX)
    images_headers, images_rows = load_sheet_rows(IMAGES_MASTER_XLSX)
    ensure_headers(master_headers, MASTER_REQUIRED, "master.xlsx")
    ensure_headers(images_headers, IMAGES_REQUIRED, "images_master.xlsx")

    work_map = defaultdict(lambda: {"witnesses": set(), "manifests": set()})

    for row in images_rows:
        work_id = clean(row.get("work_id")).zfill(3)
        witness_id = clean(row.get("witness_id"))
        m_id = clean(row.get("manifest_id"))
        if not work_id:
            continue
        if witness_id:
            work_map[work_id]["witnesses"].add(witness_id)
        if m_id:
            work_map[work_id]["manifests"].add(manifest_url(m_id))

    output = []
    for row in master_rows:
        work_id = clean(row.get("id")).zfill(3)
        if not work_id:
            continue

        title = clean(row.get("title"))
        source = clean(row.get("source"))
        date_year = clean(row.get("date_year"))
        date_month = clean(row.get("date_month"))
        date_day = clean(row.get("date_day"))
        seq = clean(row.get("seq")) or str(len(output) + 1)

        existing_default_witness = clean(row.get("default_witness"))
        existing_default_manifest = clean(row.get("default_manifest"))
        existing_manifest = clean(row.get("manifest"))

        witnesses = sorted(work_map[work_id]["witnesses"], key=sort_witness_key)
        manifests = sorted(work_map[work_id]["manifests"], key=sort_manifest_key)

        default_witness = existing_default_witness if existing_default_witness in witnesses else (witnesses[0] if witnesses else "")
        if existing_default_manifest and existing_default_manifest in manifests:
            default_manifest = existing_default_manifest
        else:
            default_manifest = ""
            if default_witness:
                suffix = f"-{default_witness}"
                for m in manifests:
                    if m.endswith(f"{suffix}.json"):
                        default_manifest = m
                        break
            if not default_manifest and manifests:
                default_manifest = manifests[0]

        manifest = existing_manifest if existing_manifest in manifests else default_manifest

        output.append({
            "seq": seq,
            "id": work_id,
            "title": title,
            "source": source,
            "date_year": date_year,
            "date_month": date_month,
            "date_day": date_day,
            "default_witness": default_witness,
            "witness_list": witnesses,
            "manifest": manifest,
            "default_manifest": default_manifest,
            "manifest_list": manifests,
            "witness_count": str(len(witnesses)),
            "has_multiple_witness": "1" if len(witnesses) > 1 else "0",
        })

    with open(MASTER_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"已更新: {MASTER_JSON}")
    print(f"共写入 {len(output)} 条记录")

if __name__ == "__main__":
    main()
