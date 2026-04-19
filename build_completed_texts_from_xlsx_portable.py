
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build text JSON files from soseki-manuscripts-text-master.xlsx.

Outputs:
- completed_texts.json        : whole-poem display for current homepage
- completed_texts_lines.json  : line-level display / research
- completed_texts_chars.json  : character-level display / research

Primary sheets:
- 主页完成テクスト
- 行文本表
- 字文本表
"""

from __future__ import annotations
import json
from pathlib import Path
import sys

try:
    import openpyxl
except ImportError:
    print("Missing dependency: openpyxl", file=sys.stderr)
    raise

REPO = Path(__file__).resolve().parent
XLSX = REPO / "soseki-manuscripts-text-master.xlsx"

WHOLE_OUTPUT = REPO / "completed_texts.json"
LINES_OUTPUT = REPO / "completed_texts_lines.json"
CHARS_OUTPUT = REPO / "completed_texts_chars.json"

WHOLE_SHEET = "主页完成テクスト"
LINES_SHEET = "行文本表"
CHARS_SHEET = "字文本表"

WHOLE_REQUIRED = ["id", "title", "source", "text", "text_kanbun", "body", "text_block"]
LINES_REQUIRED = [
    "work_id", "witness_id", "version_id", "line_no", "line_group", "line_type",
    "line_text_kanbun", "line_text_kundoku", "line_text_pair", "display_order"
]
CHARS_REQUIRED = [
    "work_id", "witness_id", "version_id", "line_no", "char_layer", "char_no",
    "char_text", "char_uid", "is_space", "is_punctuation"
]

READY_VALUES = {"", "ready", "published", "ok", "1", "true"}

def norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace("\r\n", "\n").replace("\r", "\n").strip()
    return str(v).strip()

def as_int(v, default=0):
    try:
        if v is None or v == "":
            return default
        return int(v)
    except Exception:
        return default

def as_bool(v):
    s = norm(v).lower()
    return s in {"1", "true", "yes", "y"}

def sheet_rows(ws):
    headers = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, header_map[h]).value for h in header_map.keys()}
        yield r, row, header_map

def ensure_headers(header_map, required, sheet_name):
    missing = [h for h in required if h not in header_map]
    if missing:
        raise KeyError(f"Missing headers in {sheet_name}: {', '.join(missing)}")

def load_whole(ws):
    _, _, header_map = next(sheet_rows(ws), (None, None, {}))
    ensure_headers(header_map, WHOLE_REQUIRED, WHOLE_SHEET)

    rows = []
    seen = set()

    for _, row, _ in sheet_rows(ws):
        work_id = norm(row.get("id") or row.get("work_id"))
        if not work_id:
            continue

        status = norm(row.get("status")).lower()
        if status not in READY_VALUES:
            continue

        item = {
            "id": work_id.zfill(3),
            "title": norm(row.get("title")),
            "text_block": norm(row.get("text_block") or row.get("text") or row.get("body") or row.get("text_kanbun")),
            "source": norm(row.get("source")),
            "text": norm(row.get("text") or row.get("body") or row.get("text_kanbun") or row.get("text_block")),
            "text_kanbun": norm(row.get("text_kanbun") or row.get("text") or row.get("text_block")),
            "body": norm(row.get("body") or row.get("text") or row.get("text_kanbun") or row.get("text_block")),
        }

        for extra in [
            "work_id", "witness_id", "version_id", "display_mode_default",
            "title_raw", "title_kundoku", "text_kundoku", "note"
        ]:
            val = norm(row.get(extra))
            if val:
                item[extra] = val

        if item["id"] in seen:
            raise ValueError(f"Duplicate id in {WHOLE_SHEET}: {item['id']}")
        seen.add(item["id"])
        rows.append(item)

    rows.sort(key=lambda x: x["id"])
    return rows

def load_lines(ws):
    _, _, header_map = next(sheet_rows(ws), (None, None, {}))
    ensure_headers(header_map, LINES_REQUIRED, LINES_SHEET)

    by_version = {}

    for _, row, _ in sheet_rows(ws):
        work_id = norm(row.get("work_id"))
        version_id = norm(row.get("version_id"))
        if not work_id or not version_id:
            continue

        status = norm(row.get("status")).lower()
        if status not in READY_VALUES:
            continue

        version = by_version.setdefault(version_id, {
            "id": work_id.zfill(3),
            "work_id": work_id.zfill(3),
            "witness_id": norm(row.get("witness_id")),
            "version_id": version_id,
            "title_raw": norm(row.get("title_raw")),
            "title_kundoku": norm(row.get("title_kundoku")),
            "display_mode_default": "lines",
            "lines": []
        })

        line = {
            "line_no": as_int(row.get("line_no")),
            "display_order": as_int(row.get("display_order"), as_int(row.get("line_no"))),
            "line_group": norm(row.get("line_group")),
            "line_type": norm(row.get("line_type")),
            "kanbun": norm(row.get("line_text_kanbun")),
            "kundoku": norm(row.get("line_text_kundoku")),
            "pair": norm(row.get("line_text_pair")),
            "is_visible": as_bool(row.get("is_visible")) if norm(row.get("is_visible")) else True,
            "note": norm(row.get("note")),
        }
        version["lines"].append(line)

    rows = list(by_version.values())
    for item in rows:
        item["lines"].sort(key=lambda x: (x["display_order"], x["line_no"]))
    rows.sort(key=lambda x: x["id"])
    return rows

def load_chars(ws):
    _, _, header_map = next(sheet_rows(ws), (None, None, {}))
    ensure_headers(header_map, CHARS_REQUIRED, CHARS_SHEET)

    by_version = {}
    by_line = {}

    for _, row, _ in sheet_rows(ws):
        work_id = norm(row.get("work_id"))
        version_id = norm(row.get("version_id"))
        if not work_id or not version_id:
            continue

        status = norm(row.get("status")).lower()
        if status not in READY_VALUES:
            continue

        version = by_version.setdefault(version_id, {
            "id": work_id.zfill(3),
            "work_id": work_id.zfill(3),
            "witness_id": norm(row.get("witness_id")),
            "version_id": version_id,
            "display_mode_default": "chars",
            "lines": []
        })

        line_no = as_int(row.get("line_no"))
        key = (version_id, line_no)
        if key not in by_line:
            line_obj = {"line_no": line_no, "layers": {}}
            by_line[key] = line_obj
            version["lines"].append(line_obj)

        line_obj = by_line[key]
        layer = norm(row.get("char_layer")) or "kanbun"
        layer_list = line_obj["layers"].setdefault(layer, [])
        layer_list.append({
            "char_no": as_int(row.get("char_no")),
            "char": norm(row.get("char_text")),
            "char_uid": norm(row.get("char_uid")),
            "is_space": as_bool(row.get("is_space")),
            "is_punctuation": as_bool(row.get("is_punctuation")),
            "note": norm(row.get("note")),
        })

    rows = list(by_version.values())
    for version in rows:
        version["lines"].sort(key=lambda x: x["line_no"])
        for line in version["lines"]:
            for layer, chars in list(line["layers"].items()):
                chars.sort(key=lambda x: x["char_no"])
    rows.sort(key=lambda x: x["id"])
    return rows

def write_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def main():
    if not XLSX.exists():
        raise FileNotFoundError(f"Missing workbook: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    for sheet_name in [WHOLE_SHEET, LINES_SHEET, CHARS_SHEET]:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Missing sheet: {sheet_name}")

    whole_rows = load_whole(wb[WHOLE_SHEET])
    line_rows = load_lines(wb[LINES_SHEET])
    char_rows = load_chars(wb[CHARS_SHEET])

    write_json(WHOLE_OUTPUT, whole_rows)
    write_json(LINES_OUTPUT, line_rows)
    write_json(CHARS_OUTPUT, char_rows)

    print(f"Wrote {WHOLE_OUTPUT.name} ({len(whole_rows)} items)")
    print(f"Wrote {LINES_OUTPUT.name} ({len(line_rows)} versions)")
    print(f"Wrote {CHARS_OUTPUT.name} ({len(char_rows)} versions)")

if __name__ == "__main__":
    main()
