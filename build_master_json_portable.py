#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import json
import sys
from pathlib import Path
from collections import defaultdict
import openpyxl

REPO_ROOT = Path(__file__).resolve().parent
MASTER_XLSX = REPO_ROOT / "master.xlsx"
IMAGES_MASTER_XLSX = REPO_ROOT / "images_master.xlsx"
MASTER_JSON = REPO_ROOT / "master.json"
BASE_URL = "https://scholarbo.github.io/soseki-manuscripts/manifests"

MASTER_REQUIRED = ["seq", "id", "title", "source", "date_year", "date_month", "date_day"]
IMAGES_REQUIRED = ["work_id", "witness_id", "manifest_id"]

def clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def load_sheet_rows(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return headers, rows

def ensure_headers(headers, required, label):
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"{label} 缺少必要列: {missing}")

def main():
    if not MASTER_XLSX.exists():
        raise FileNotFoundError(f"缺少文件: {MASTER_XLSX}")
    if not IMAGES_MASTER_XLSX.exists():
        raise FileNotFoundError(f"缺少文件: {IMAGES_MASTER_XLSX}")

    master_headers, master_rows = load_sheet_rows(MASTER_XLSX)
    images_headers, images_rows = load_sheet_rows(IMAGES_MASTER_XLSX)

    ensure_headers(master_headers, MASTER_REQUIRED, "master.xlsx")
    ensure_headers(images_headers, IMAGES_REQUIRED, "images_master.xlsx")

    by_work = defaultdict(list)
    for row in images_rows:
        work_id = clean(row.get("work_id"))
        witness_id = clean(row.get("witness_id")) or "w01"
        manifest_id = clean(row.get("manifest_id"))
        if not work_id or not manifest_id:
            continue
        by_work[work_id].append({
            "witness_id": witness_id,
            "manifest_id": manifest_id,
        })

    result = []
    for row in master_rows:
        work_id = clean(row.get("id"))
        if not work_id:
            continue

        manifests = by_work.get(work_id, [])
        manifest_list = []
        witness_list = []

        for m in manifests:
            witness_id = m["witness_id"]
            manifest_id = m["manifest_id"]
            manifest_file = f"{manifest_id}.json" if not manifest_id.endswith(".json") else manifest_id
            manifest_url = f"{BASE_URL}/{manifest_file}"
            manifest_list.append({
                "witness_id": witness_id,
                "manifest_id": manifest_id.removesuffix(".json"),
                "manifest_url": manifest_url,
            })
            witness_list.append(witness_id)

        default_manifest = manifest_list[0]["manifest_url"] if manifest_list else ""
        default_witness = manifest_list[0]["witness_id"] if manifest_list else ""

        out = {k: clean(v) for k, v in row.items()}
        out["default_witness"] = default_witness
        out["default_manifest"] = default_manifest
        out["manifest_list"] = manifest_list
        out["witness_list"] = witness_list
        result.append(out)

    with open(MASTER_JSON, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"已更新: {MASTER_JSON}")
    print(f"共写入 {len(result)} 条记录")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
