#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent
IMAGES_DIR = REPO_ROOT / "images"
IMAGES_MASTER_XLSX = REPO_ROOT / "images_master.xlsx"
MANIFEST_DIR = REPO_ROOT / "manifests"
MASTER_JSON = REPO_ROOT / "master.json"
COMPLETED_TEXTS_JSON = REPO_ROOT / "completed_texts.json"

BASE_URL = "https://scholarbo.github.io/soseki-manuscripts"
SUPPORTED_IMAGE_FORMATS = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".tif": "image/tiff",
    ".tiff": "image/tiff",
}


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()



def load_json(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)



def image_size(path: Path):
    with Image.open(path) as im:
        return im.size



def image_format_for(path: Path) -> str:
    return SUPPORTED_IMAGE_FORMATS.get(path.suffix.lower(), "image/jpeg")



def build_metadata(master_entry: dict, witness_id: str):
    md = []

    def add(label, value):
        value = clean(value)
        if value:
            md.append({"label": {"ja": [label]}, "value": {"ja": [value]}})

    add("作品ID", master_entry.get("id") or master_entry.get("work_id"))
    add("作品名", master_entry.get("title"))
    add("Witness", witness_id)
    add("source", master_entry.get("source"))
    add("date_year", master_entry.get("date_year"))
    add("date_month", master_entry.get("date_month"))
    add("date_day", master_entry.get("date_day"))
    return md



def sheet_rows_from_images_master(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [clean(c.value) for c in ws[1]]
    required = {
        "work_id",
        "witness_id",
        "page",
        "image_role",
        "source_path",
        "iiif_path",
        "label",
        "manifest_id",
        "canvas_id",
    }
    missing = [h for h in sorted(required) if h not in headers]
    if missing:
        raise KeyError(f"images_master.xlsx 缺少必要列: {', '.join(missing)}")

    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = {h: row[i] if i < len(row) else None for h, i in idx.items()}
        work_id = clean(values.get("work_id"))
        if not work_id:
            continue
        rows.append({k: values.get(k) for k in headers})
    return rows



def resolve_relative_path(row: dict) -> str:
    image_role = clean(row.get("image_role")).lower()
    iiif_path = clean(row.get("iiif_path"))
    source_path = clean(row.get("source_path"))
    web_path = clean(row.get("web_path"))

    if image_role == "iiif" and iiif_path:
        return iiif_path
    if source_path:
        return source_path
    if iiif_path:
        return iiif_path
    if web_path.startswith(BASE_URL + "/"):
        return web_path.replace(BASE_URL + "/", "", 1)
    if web_path:
        return web_path
    return ""



def resolve_public_url(rel_or_url: str) -> str:
    rel_or_url = clean(rel_or_url)
    if not rel_or_url:
        return ""
    if rel_or_url.startswith("http://") or rel_or_url.startswith("https://"):
        return rel_or_url
    return f"{BASE_URL}/{rel_or_url.lstrip('/')}"



def completed_summary(text_entry: dict) -> list[str]:
    original_text = (
        clean(text_entry.get("original"))
        or clean(text_entry.get("original_text"))
        or clean(text_entry.get("text_kanbun"))
        or clean(text_entry.get("text"))
        or clean(text_entry.get("body"))
    )
    kundoku_text = (
        clean(text_entry.get("kundoku"))
        or clean(text_entry.get("kundoku_text"))
    )

    lines: list[str] = []
    if original_text:
        lines.append("【原文】")
        lines.append(original_text)
    if kundoku_text:
        if lines:
            lines.append("")
        lines.append("【訓読】")
        lines.append(kundoku_text)
    return lines



def main():
    if not IMAGES_DIR.exists():
        raise FileNotFoundError(f"缺少目录: {IMAGES_DIR}")
    if not IMAGES_MASTER_XLSX.exists():
        raise FileNotFoundError(f"缺少文件: {IMAGES_MASTER_XLSX}")
    if not MASTER_JSON.exists():
        raise FileNotFoundError(f"缺少文件: {MASTER_JSON}")
    if not COMPLETED_TEXTS_JSON.exists():
        raise FileNotFoundError(f"缺少文件: {COMPLETED_TEXTS_JSON}")

    MANIFEST_DIR.mkdir(parents=True, exist_ok=True)

    image_rows = sheet_rows_from_images_master(IMAGES_MASTER_XLSX)
    print(f"images_master.xlsx 数据行数: {len(image_rows)}")

    master_rows = load_json(MASTER_JSON)
    completed_rows = load_json(COMPLETED_TEXTS_JSON)
    print(f"master.json 条目数: {len(master_rows)}")
    print(f"completed_texts.json 条目数: {len(completed_rows)}")

    master_map = {clean(x.get("id") or x.get("work_id")): x for x in master_rows}
    text_map = {clean(x.get("id") or x.get("work_id")): x for x in completed_rows}

    grouped: dict[str, list[dict]] = defaultdict(list)
    warnings: list[str] = []

    for row in image_rows:
        role = clean(row.get("image_role")).lower()
        if role and role != "iiif":
            continue

        work_id = clean(row.get("work_id"))
        witness_id = clean(row.get("witness_id")).lower() or "w01"
        manifest_id = clean(row.get("manifest_id")) or f"manifest-{work_id}-{witness_id}"
        rel_path = resolve_relative_path(row)
        if not rel_path:
            warnings.append(f"跳过 {work_id}/{witness_id}/p{row.get('page')}: 缺少可用图片路径")
            continue

        local_path = REPO_ROOT / rel_path
        if not local_path.exists():
            warnings.append(f"跳过 {work_id}/{witness_id}/p{row.get('page')}: 图片不存在 {rel_path}")
            continue

        try:
            page_num = int(row.get("page"))
        except Exception:
            warnings.append(f"跳过 {work_id}/{witness_id}: 非法页码 {row.get('page')}")
            continue

        width, height = image_size(local_path)
        public_url = resolve_public_url(rel_path)
        canvas_id = clean(row.get("canvas_id")) or f"canvas-{work_id}-{witness_id}" + ("" if page_num == 1 else f"-p{page_num}")
        grouped[manifest_id].append(
            {
                "work_id": work_id,
                "witness_id": witness_id,
                "page": page_num,
                "label": clean(row.get("label")),
                "manifest_id": manifest_id,
                "canvas_id": canvas_id,
                "rel_path": rel_path,
                "local_path": local_path,
                "public_url": public_url,
                "width": width,
                "height": height,
                "format": image_format_for(local_path),
            }
        )

    count = 0
    for manifest_id in sorted(grouped.keys()):
        pages = sorted(grouped[manifest_id], key=lambda x: (x["page"], x["canvas_id"]))
        first = pages[0]
        work_id = first["work_id"]
        witness_id = first["witness_id"]
        manifest_url = f"{BASE_URL}/manifests/{manifest_id}.json"

        master_entry = master_map.get(work_id, {})
        text_entry = text_map.get(work_id, {})
        manifest_label = clean(master_entry.get("title")) or clean(text_entry.get("title")) or manifest_id

        items = []
        for page in pages:
            canvas_url = f"{manifest_url}/{page['canvas_id']}"
            anno_page_id = f"{canvas_url}/page"
            anno_id = f"{canvas_url}/annotation"
            canvas_label = page["label"] or f"{manifest_label} 第{page['page']}頁"
            items.append(
                {
                    "id": canvas_url,
                    "type": "Canvas",
                    "width": page["width"],
                    "height": page["height"],
                    "label": {"ja": [canvas_label]},
                    "items": [
                        {
                            "id": anno_page_id,
                            "type": "AnnotationPage",
                            "items": [
                                {
                                    "id": anno_id,
                                    "type": "Annotation",
                                    "motivation": "painting",
                                    "target": canvas_url,
                                    "body": {
                                        "id": page["public_url"],
                                        "type": "Image",
                                        "format": page["format"],
                                        "width": page["width"],
                                        "height": page["height"],
                                    },
                                }
                            ],
                        }
                    ],
                }
            )

        manifest = {
            "@context": "http://iiif.io/api/presentation/3/context.json",
            "id": manifest_url,
            "type": "Manifest",
            "label": {"ja": [manifest_label]},
            "metadata": build_metadata(master_entry, witness_id),
            "items": items,
        }

        summary_lines = completed_summary(text_entry)
        if summary_lines:
            manifest["summary"] = {"ja": ["\n".join(summary_lines)]}

        out_path = MANIFEST_DIR / f"{manifest_id}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)

        print(f"已生成: {out_path} ({len(items)} page(s))")
        count += 1

    print(f"共生成 {count} 个 manifest")
    if warnings:
        print("\n警告：")
        for msg in warnings:
            print(f"- {msg}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
